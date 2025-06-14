import os
import smtplib
import sqlite3
from email import encoders
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
from openpyxl.workbook import Workbook

from twilio_conf import send_sms

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooser
from kivy.uix.popup import Popup
import pyshark


def calculate_real_data_rate(pcap_file, ip):
    print("Calculate real data rate method", ip)
    """
    Calculate the real data rate (application-layer payload) for TCP and UDP traffic.

    Args:
        pcap_file (str): Path to the .pcap file.

    Returns:
        dict: Contains total data size, duration, and data rate.
    """
    capture = pyshark.FileCapture(pcap_file)

    total_data_size = 0
    start_time = None
    end_time = None
    target_ip = ip[1]
    for packet in capture:
        try:
            if 'IP' in packet:

                ip_layer = packet.ip
                # Filter packets based on the target IP address
                if ip_layer.src == target_ip or ip_layer.dst == target_ip:

                    if 'TCP' in packet:

                        tcp_layer = packet.tcp
                        if hasattr(tcp_layer, 'len') and int(tcp_layer.len) > 0:
                            total_data_size += int(tcp_layer.len)
                    elif 'UDP' in packet:

                        udp_layer = packet.udp
                        if hasattr(udp_layer, 'length') and int(udp_layer.length) > 8:
                            total_data_size += int(udp_layer.length) - 8
                    # Update start and end timestamps
                    packet_time = float(packet.sniff_timestamp)
                    if start_time is None:
                        start_time = packet_time
                    end_time = packet_time
        except AttributeError:
            # Ignore packets that don't have relevant attributes
            continue

    capture.close()

    if start_time and end_time:
        duration = end_time - start_time
        total_data_bits = total_data_size * 8
        data_rate_mbps = total_data_bits / (duration * 1_000_000)
        #save_datarate_in_db(data_rate_mbps)
        # return {
        #     "total_payload_size_bytes": total_data_size,
        #     "duration_seconds": duration,
        #     "data_rate_mbps": data_rate_mbps
        # }
        return data_rate_mbps
    else:
        return {"error": "No relevant TCP/UDP packets found in the capture."}

def save_datarate_in_db(data_rate_mbps, filename):
    print("data rate type", type(data_rate_mbps))
    """Save packet capture info in the database."""
    try:
        conn = sqlite3.connect("../db/file_storage.db")

        cursor = conn.cursor()

        cursor.execute("UPDATE file_storage SET DATA_RATE = ? WHERE filename = ?",
                       (data_rate_mbps, filename))
        conn.commit()
        conn.close()
    except sqlite3.Error as e:
        print('error',e)


def compare_with_average_data_rate(current_data_rate, ip):
    print("compare data rates")

    output_file = os.path.abspath("file_storage_report.xlsx")
    """
    Compare the current data rate with the average of the last 7 data rate values in the database.

    Args:
        db_path (str): Path to the SQLite database file.
        table_name (str): Name of the table containing the data rate column.
        current_data_rate (float): The data rate of the currently captured file.

    Returns:
        dict: Contains the average data rate and comparison result.
    """

    db_path = "../db/file_storage.db"

    try:
        # Connect to the database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Fetch the last 7 entries from the DATA_RATE column
        cursor.execute(f"SELECT DATA_RATE FROM Files WHERE IP = ? ORDER BY time DESC LIMIT 8", (ip,))
        data_rate_entries = cursor.fetchall()
        #print("data rate entries", data_rate_entries)

        # Exclude the current data rate (the last entry fetched is the current one)
        previous_data_rates = [entry[0] for entry in data_rate_entries[:-1] if entry[0] is not None]

        # Check if there are enough entries
        if len(previous_data_rates) < 7:
            print ("error: Not enough entries in the database to calculate average")
            return

        average_data_rate = sum(previous_data_rates) / len(previous_data_rates)
        print("average data rate", average_data_rate)
        print("current data rate", current_data_rate)

        message = f"""Dear Team,

        An abnormal data rate has been detected in the system.

        Details:
        - Current Data Rate: {current_data_rate} Mbps
        - Average Data Rate: {average_data_rate} Mbps
        - Threshold (2x Average): {2 * average_data_rate} Mbps

        This indicates a potential issue requiring immediate attention.

        Please investigate the source of the anomaly.

        Best regards,
        VAP IoT Application Scanner"""

        print("data rate comparison",current_data_rate, 2 * average_data_rate)
        # Compare the current data rate with the average
        if current_data_rate > 2 * average_data_rate:
            print("abnormal data rate")
            exported_file = export_to_excel(db_path, output_file, ip)
            send_mail(message, ip, exported_file)
            send_sms(message)



        # Close the database connection
        conn.close()


    except sqlite3.Error as e:
        print("Database error:", e)
        #return {"error": str(e)}



def send_mail(message, excel_file_path, exported_file):
    """Send an email with the specified message and attach the Excel report."""
    print("Triggering mail")

    # Email credentials
    email = "your_mail@gmail.com"

    password = "your_password"

    if not email or not password:
        print("Environment variables for email credentials are not set.")
        return

    try:
        # Create the email message
        msg = MIMEMultipart()
        msg["From"] = email
        msg["To"] = email  # Sending to the same email for now, can be changed to other recipients
        msg["Subject"] = "Abnormal Data Rate Detected - IoT Report"

        # Attach the main message
        msg.attach(MIMEText(message, "plain"))

        # Attach the Excel file if it exists
        if os.path.exists(exported_file):
            with open(exported_file, "rb") as file:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(exported_file)}",
                )
                msg.attach(part)
        else:
            print("Error: Excel file not found for attachment.")

        # Set up the SMTP server
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(email, password)
        server.send_message(msg)
        server.quit()

        print("Email sent successfully with the attachment!")
    except Exception as e:
        print(f"Failed to send email: {e}")

def export_to_excel(db_path, output_file, ip):
    """Export database entries for a specific IP address to an Excel file."""
    try:
        # Connect to the database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Query to fetch records for the specific IP
        cursor.execute("SELECT * FROM Files WHERE IP = ?", (ip,))
        records = cursor.fetchall()

        # Get the column names
        column_names = [description[0] for description in cursor.description]

        # Close the database connection
        conn.close()

        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Capture History"

        # Add column headers
        for col_num, column_name in enumerate(column_names, start=1):
            sheet.cell(row=1, column=col_num, value=column_name)

        # Add rows of data
        for row_num, record in enumerate(records, start=2):
            for col_num, value in enumerate(record, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Save the Excel file
        workbook.save(output_file)
        print(f"Excel file created: {output_file}")
        return output_file

    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return None
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return None



class IDSApp(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", **kwargs)
        self.file_label = Label(text="Select a PCAP file to analyze", size_hint=(1, 0.2))
        self.add_widget(self.file_label)

        self.file_chooser_button = Button(text="Choose File", size_hint=(1, 0.2))
        self.file_chooser_button.bind(on_press=self.open_file_chooser)
        self.add_widget(self.file_chooser_button)

        self.result_label = Label(text="", size_hint=(1, 0.6))
        self.add_widget(self.result_label)

    def open_file_chooser(self, instance):
        file_chooser = FileChooser(path=os.getcwd(), filters=["*.pcap"])
        popup = Popup(title="Select a PCAP file", content=file_chooser, size_hint=(0.9, 0.9))
        file_chooser.bind(on_selection=self.load_file)
        popup.open()

    def load_file(self, instance, selection):
        if selection:
            pcap_file = selection[0]
            self.file_label.text = f"Selected File: {os.path.basename(pcap_file)}"
            result = calculate_real_data_rate(pcap_file)

            if "error" in result:
                self.result_label.text = result["error"]
            else:
                self.result_label.text = (
                    f"Total Payload Size: {result['total_payload_size_bytes']} bytes\n"
                    f"Duration: {result['duration_seconds']:.2f} seconds\n"
                    f"Data Rate: {result['data_rate_mbps']:.2f} Mbps"
                )


class IDSAppMain(App):
    def build(self):
        return IDSApp()


if __name__ == "__main__":
    IDSAppMain().run()
